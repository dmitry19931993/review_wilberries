import openpyxl
import requests
import os
from dotenv import load_dotenv




def take_sku(file_path):
    sku_list = []
    wookbook = openpyxl.load_workbook(f"{file_path}")
    worksheet = wookbook.active
    worksheet.append(['dsfbdfb'])
    wookbook.save(f"{file_path}")
    for i in worksheet.values:
        sku_list.append(i[0])
    return sku_list

def id_feedback(sku):
    link = f'https://card.wb.ru/cards/v2/detail?nm={sku}'
    my_req = requests.get(link)
    if my_req.status_code == 200:
        data = my_req.json()
        if len(data['data']['products']) > 0:
            id_feedback = data['data']['products'][0]['root']
            name = data['data']['products'][0]['name']
            raiting = data['data']['products'][0]['supplierRating']
        else:
            return None
    else:
        return None
    return id_feedback, name, raiting

def check_feedback_link(id_feed):
    link = f'https://feedbacks1.wb.ru/feedbacks/v1/{id_feed}'
    link_2 = f'https://feedbacks2.wb.ru/feedbacks/v1/{id_feed}'
    my_req = requests.get(link)
    feedback_list = []
    if my_req.status_code == 200:
        feedbacks = my_req.json()
        if feedbacks['feedbacks']:
            feedbacks_list = feedbacks['feedbacks']
        else:
            my_req = requests.get(link_2)
            if my_req.status_code == 200:
                feedbacks = my_req.json()
                if feedbacks['feedbacks']:
                    feedbacks_list = feedbacks['feedbacks']
    return feedbacks_list


def take_feedback(id_feed, feedback_pk_list):
    feedbacks_list = check_feedback_link(id_feed)
    feedbacks_data_list = []
    if feedbacks_list:
        for i in feedbacks_list:
            if int(i['productValuation']) <= 4:
                valuation = i['productValuation']
                text = i['text']
                feedback_pk = i['id']
                if feedback_pk not in feedback_pk_list:
                    feedbacks_data_list.append([valuation, text])
                    feedback_pk_list.append(feedback_pk)
        return feedbacks_data_list


def get_feedback_pk(file_path):
    feedback_pk_list = []
    wookbook = openpyxl.load_workbook(f"{file_path}")
    worksheet = wookbook.active
    for i in worksheet.values:
        feedback_pk_list.append(i[0])
    return feedback_pk_list

def save_feedback_pk(fedback_pk_list, file_path):
    wookbook = openpyxl.load_workbook(f"{file_path}")
    worksheet = wookbook.active
    for i in fedback_pk_list:
        worksheet.append([i])
    wookbook.save(f"{file_path}")

def send_message(message):
    load_dotenv()
    TOKEN = os.getenv("API_KEY")
    chat_id = os.getenv("CHAT_ID")
    url = f"https://api.telegram.org/bot{TOKEN}/sendMessage?chat_id={chat_id}&text={message}"
    my_req = requests.get(url)
    if my_req.status_code == 200:
        print('успешно отправлено')
    else:
        print('ошибка отправления')





def main():
    sku_list = take_sku(file_path="SKU.xlsx")
    feedback_pk_list = get_feedback_pk(file_path='feedback_pk.xlsx')
    for sku in sku_list:
        data = id_feedback(sku)
        if data:
            id_feed = data[0]
            name = data[1]
            raiting = data[2]
            feedbacks_data_list = take_feedback(id_feed, feedback_pk_list)
            if feedbacks_data_list:
                for feedback in feedbacks_data_list:
                    message = (f'негативный отзыв/{name}/{sku}/{feedback[0]}/{feedback[1]}/{raiting}')
                    send_message(message)

    save_feedback_pk(feedback_pk_list, file_path='feedback_pk.xlsx')




if __name__ == "__main__":
    main()